import sys
import os
import numpy as np
import pandas as pd
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                             QPushButton, QLabel, QSlider, QFileDialog, QListWidget, 
                             QSplitter, QMessageBox, QScrollArea, QGroupBox, QSpinBox)
from PyQt5.QtCore import Qt, QDir
from PyQt5.QtGui import QPixmap, QImage, QPainter, QPen
import pydicom
from pydicom.pixel_data_handlers.util import apply_voi_lut
from pydicom.dataelem import DataElement
from pydicom.multival import MultiValue
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

class DicomViewer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.src_folder = ""
        self.dst_folder = ""
        self.current_angle = 0
        self.angle_records = {}  # 存储每个子文件夹的角度记录
        self.subfolders = []  # 存储所有子文件夹路径
        self.current_subfolder_index = -1  # 当前选中的子文件夹索引
        self.dicom_files = []  # 当前子文件夹中的所有DICOM文件
        self.current_dicom_index = -1  # 当前显示的DICOM文件索引
        
        # 窗宽窗位参数
        self.window_width = 400
        self.window_level = 40
        self.original_image = None  # 存储原始图像数据
        self.original_pixel_array = None  # 存储原始像素数据
        
        self.initUI()
        
    def initUI(self):
        self.setWindowTitle('头颅CT图像角度矫正工具')
        self.setGeometry(100, 100, 1400, 900)
        
        # 中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主布局
        main_layout = QHBoxLayout(central_widget)
        
        # 左侧面板 - 文件夹和文件列表
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_panel.setMaximumWidth(350)
        
        # 源文件夹选择
        src_btn = QPushButton('选择源文件夹')
        src_btn.clicked.connect(self.select_src_folder)
        left_layout.addWidget(src_btn)
        
        self.src_label = QLabel('未选择源文件夹')
        self.src_label.setWordWrap(True)
        left_layout.addWidget(self.src_label)
        
        # 目标文件夹选择
        dst_btn = QPushButton('选择目标文件夹')
        dst_btn.clicked.connect(self.select_dst_folder)
        left_layout.addWidget(dst_btn)
        
        self.dst_label = QLabel('未选择目标文件夹')
        self.dst_label.setWordWrap(True)
        left_layout.addWidget(self.dst_label)
        
        # 子文件夹列表
        left_layout.addWidget(QLabel('子文件夹:'))
        self.subfolder_list = QListWidget()
        self.subfolder_list.currentRowChanged.connect(self.subfolder_selected)
        left_layout.addWidget(self.subfolder_list)
        
        # DICOM文件列表
        left_layout.addWidget(QLabel('DICOM文件:'))
        self.dicom_list = QListWidget()
        self.dicom_list.currentRowChanged.connect(self.dicom_selected)
        left_layout.addWidget(self.dicom_list)
        
        # 右侧面板 - 图像显示和控制
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        # 图像显示区域
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMinimumSize(500, 500)
        self.image_label.setText('请选择DICOM文件查看图像')
        
        scroll_area = QScrollArea()
        scroll_area.setWidget(self.image_label)
        scroll_area.setWidgetResizable(True)
        right_layout.addWidget(scroll_area, 3)
        
        # 控制区域
        control_layout = QHBoxLayout()
        
        # 角度控制组
        angle_group = QGroupBox("角度控制")
        angle_layout = QVBoxLayout(angle_group)
        
        angle_slider_layout = QHBoxLayout()
        angle_slider_layout.addWidget(QLabel('旋转角度:'))
        
        self.angle_slider = QSlider(Qt.Horizontal)
        self.angle_slider.setRange(-180, 180)
        self.angle_slider.setValue(0)
        self.angle_slider.valueChanged.connect(self.angle_changed)
        angle_slider_layout.addWidget(self.angle_slider)
        
        self.angle_value = QLabel('0°')
        angle_slider_layout.addWidget(self.angle_value)
        
        angle_layout.addLayout(angle_slider_layout)
        control_layout.addWidget(angle_group)
        
        # 窗宽窗位控制组
        ww_wl_group = QGroupBox("窗宽窗位控制")
        ww_wl_layout = QVBoxLayout(ww_wl_group)
        
        # 窗宽控制
        ww_layout = QHBoxLayout()
        ww_layout.addWidget(QLabel('窗宽:'))
        self.ww_slider = QSlider(Qt.Horizontal)
        self.ww_slider.setRange(1, 2000)
        self.ww_slider.setValue(self.window_width)
        self.ww_slider.valueChanged.connect(self.ww_changed)
        ww_layout.addWidget(self.ww_slider)
        
        self.ww_value = QLabel(f'{self.window_width}')
        ww_layout.addWidget(self.ww_value)
        ww_wl_layout.addLayout(ww_layout)
        
        # 窗位控制
        wl_layout = QHBoxLayout()
        wl_layout.addWidget(QLabel('窗位:'))
        self.wl_slider = QSlider(Qt.Horizontal)
        self.wl_slider.setRange(-1000, 1000)
        self.wl_slider.setValue(self.window_level)
        self.wl_slider.valueChanged.connect(self.wl_changed)
        wl_layout.addWidget(self.wl_slider)
        
        self.wl_value = QLabel(f'{self.window_level}')
        wl_layout.addWidget(self.wl_value)
        ww_wl_layout.addLayout(wl_layout)
        
        control_layout.addWidget(ww_wl_group)
        right_layout.addLayout(control_layout, 1)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        
        prev_btn = QPushButton('上一张')
        prev_btn.clicked.connect(self.prev_image)
        btn_layout.addWidget(prev_btn)
        
        next_btn = QPushButton('下一张')
        next_btn.clicked.connect(self.next_image)
        btn_layout.addWidget(next_btn)
        
        reset_ww_wl_btn = QPushButton('重置窗宽窗位')
        reset_ww_wl_btn.clicked.connect(self.reset_ww_wl)
        btn_layout.addWidget(reset_ww_wl_btn)
        
        apply_btn = QPushButton('应用角度')
        apply_btn.clicked.connect(self.apply_angle)
        btn_layout.addWidget(apply_btn)
        
        save_btn = QPushButton('保存到Excel')
        save_btn.clicked.connect(self.save_to_excel)
        btn_layout.addWidget(save_btn)
        
        right_layout.addLayout(btn_layout)
        
        # 分割左右面板
        splitter = QSplitter(Qt.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 1100])
        
        main_layout.addWidget(splitter)
        
    def select_src_folder(self):
        folder = QFileDialog.getExistingDirectory(self, '选择源文件夹')
        if folder:
            self.src_folder = folder
            self.src_label.setText(f'源文件夹: {folder}')
            self.scan_subfolders()
            
    def select_dst_folder(self):
        folder = QFileDialog.getExistingDirectory(self, '选择目标文件夹')
        if folder:
            self.dst_folder = folder
            self.dst_label.setText(f'目标文件夹: {folder}')
            
    def scan_subfolders(self):
        """扫描源文件夹中的所有子文件夹"""
        self.subfolders = []
        self.subfolder_list.clear()
        
        if not self.src_folder:
            return
            
        # 遍历源文件夹中的所有子文件夹
        for root, dirs, files in os.walk(self.src_folder):
            # 检查是否有DICOM文件
            has_dicom = any(file.lower().endswith('.dcm') for file in files)
            if has_dicom:
                self.subfolders.append(root)
                # 显示相对于源文件夹的路径
                rel_path = os.path.relpath(root, self.src_folder)
                self.subfolder_list.addItem(rel_path)
                
    def subfolder_selected(self, index):
        """当选择子文件夹时"""
        if index < 0 or index >= len(self.subfolders):
            return
            
        self.current_subfolder_index = index
        subfolder_path = self.subfolders[index]
        
        # 扫描该子文件夹中的所有DICOM文件
        self.dicom_files = []
        self.dicom_list.clear()
        
        for file in os.listdir(subfolder_path):
            if file.lower().endswith('.dcm'):
                self.dicom_files.append(os.path.join(subfolder_path, file))
                self.dicom_list.addItem(file)
                
        # 如果有DICOM文件，选择第一个
        if self.dicom_files:
            self.dicom_list.setCurrentRow(0)
            
        # 如果该子文件夹已有角度记录，更新滑块
        rel_path = os.path.relpath(subfolder_path, self.src_folder)
        if rel_path in self.angle_records:
            angle = self.angle_records[rel_path]
            self.angle_slider.setValue(angle)
            self.angle_value.setText(f'{angle}°')
            self.current_angle = angle
            
    def dicom_selected(self, index):
        """当选择DICOM文件时"""
        if index < 0 or index >= len(self.dicom_files):
            return
            
        self.current_dicom_index = index
        dicom_path = self.dicom_files[index]
        
        # 加载并显示DICOM图像
        self.display_dicom_image(dicom_path)
        
    def display_dicom_image(self, dicom_path):
        """显示DICOM图像"""
        try:
            # 读取DICOM文件
            ds = pydicom.dcmread(dicom_path)
            
            # 获取像素数据
            self.original_pixel_array = ds.pixel_array
            
            # 尝试获取DICOM文件中的窗宽窗位信息
            if hasattr(ds, 'WindowWidth') and hasattr(ds, 'WindowCenter'):
                # 处理窗宽
                ww = ds.WindowWidth
                if isinstance(ww, MultiValue):
                    ww = float(ww[0])
                elif hasattr(ww, 'value'):
                    ww = float(ww.value)
                else:
                    ww = float(ww)
                
                # 处理窗位
                wl = ds.WindowCenter
                if isinstance(wl, MultiValue):
                    wl = float(wl[0])
                elif hasattr(wl, 'value'):
                    wl = float(wl.value)
                else:
                    wl = float(wl)
                
                self.window_width = int(ww)
                self.window_level = int(wl)
                
                # 更新滑块和标签
                self.ww_slider.setValue(self.window_width)
                self.wl_slider.setValue(self.window_level)
                self.ww_value.setText(f'{self.window_width}')
                self.wl_value.setText(f'{self.window_level}')
            
                # 应用窗宽窗位
                img = self.apply_window_level(self.original_pixel_array, self.window_width, self.window_level)
                
                # 存储处理后的图像数据
                self.original_image = img
            else:
                self.original_image = self.original_pixel_array
            # 应用当前角度旋转
            if self.current_angle != 0:
                img = self.rotate_image(img, self.current_angle)
                
            # 转换为QImage并显示
            height, width = img.shape
            bytes_per_line = width
            q_img = QImage(img.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
            pixmap = QPixmap.fromImage(q_img)
            # 创建一个新的QPixmap用于绘制虚线
            pixmap_with_lines = QPixmap(pixmap.size())
            pixmap_with_lines.fill(Qt.transparent)  # 透明背景

            # 使用QPainter在pixmap上绘制
            painter = QPainter(pixmap)

            # 绘制原始图像
            painter.drawPixmap(0, 0, pixmap)

            # 设置画笔为虚线
            pen = QPen(Qt.red)  # 红色虚线，你可以改成其他颜色
            pen.setStyle(Qt.DashLine)
            pen.setWidth(2)
            painter.setPen(pen)

            # 绘制中间的垂直虚线
            center_x = pixmap.width() // 2
            painter.drawLine(center_x, 0, center_x, pixmap.height())

            painter.end()
            # 缩放以适应标签大小
            label_size = self.image_label.size()
            scaled_pixmap = pixmap.scaled(label_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.image_label.setPixmap(scaled_pixmap)
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'无法读取DICOM文件: {str(e)}')
            import traceback
            traceback.print_exc()
            
    def apply_window_level(self, image, window_width, window_level):
        """应用窗宽窗位到图像"""
        # 首先将图像转换为浮点数类型以避免溢出
        image_float = image.astype(np.float32)
        
        # 计算窗宽窗位的上下限
        window_min = window_level - window_width / 2.0
        window_max = window_level + window_width / 2.0
        img_max = image.max()
        img_min = image.min()
        # 应用窗宽窗位
        #windowed = np.clip(image_float, window_min, window_max)
        windowed = np.clip(image_float, img_min, img_max)
        # 归一化到0-255
        #window_range = window_max - window_min
        window_range = img_max - img_min
        if window_range > 0:
            #windowed = (windowed - window_min) / window_range * 255.0
            windowed = (windowed - img_min) / window_range * 255.0
        else:
            # 避免除以零
            windowed = np.zeros_like(windowed)
        
        # 转换回uint8
        windowed = windowed.astype(np.uint8)
        
        return windowed
            
    def rotate_image(self, image, angle):
        """旋转图像"""
        if angle == 0:
            return image
            
        # 获取图像中心
        (h, w) = image.shape[:2]
        center = (w // 2, h // 2)
        
        # 计算旋转矩阵
        M = cv2.getRotationMatrix2D(center, angle, 1.0)
        
        # 执行旋转
        rotated = cv2.warpAffine(image, M, (w, h))
        
        return rotated
        
    def angle_changed(self, value):
        """当角度滑块变化时"""
        self.current_angle = value
        self.angle_value.setText(f'{value}°')
        
        # 如果当前有显示的图像，重新显示应用旋转后的图像
        if self.current_dicom_index >= 0 and self.original_image is not None:
            # 应用旋转
            if self.current_angle != 0:
                img = self.rotate_image(self.original_image, self.current_angle)
            else:
                img = self.original_image
                
            # 显示图像
            height, width = img.shape
            bytes_per_line = width
            q_img = QImage(img.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
            pixmap = QPixmap.fromImage(q_img)
            # 创建一个新的QPixmap用于绘制虚线
            pixmap_with_lines = QPixmap(pixmap.size())
            pixmap_with_lines.fill(Qt.transparent)  # 透明背景

            # 使用QPainter在pixmap上绘制
            painter = QPainter(pixmap)

            # 绘制原始图像
            painter.drawPixmap(0, 0, pixmap)

            # 设置画笔为虚线
            pen = QPen(Qt.red)  # 红色虚线，你可以改成其他颜色
            pen.setStyle(Qt.DashLine)
            pen.setWidth(2)
            painter.setPen(pen)

            # 绘制中间的垂直虚线
            center_x = pixmap.width() // 2
            painter.drawLine(center_x, 0, center_x, pixmap.height())

            painter.end()
            label_size = self.image_label.size()
            scaled_pixmap = pixmap.scaled(label_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.image_label.setPixmap(scaled_pixmap)
            
    def ww_changed(self, value):
        """当窗宽滑块变化时"""
        self.window_width = value
        self.ww_value.setText(f'{value}')
        
        # 如果当前有显示的图像，重新显示应用窗宽窗位后的图像
        if self.current_dicom_index >= 0 and self.original_pixel_array is not None:
            self.display_dicom_image(self.dicom_files[self.current_dicom_index])
            
    def wl_changed(self, value):
        """当窗位滑块变化时"""
        self.window_level = value
        self.wl_value.setText(f'{value}')
        
        # 如果当前有显示的图像，重新显示应用窗宽窗位后的图像
        if self.current_dicom_index >= 0 and self.original_pixel_array is not None:
            self.display_dicom_image(self.dicom_files[self.current_dicom_index])
            
    def reset_ww_wl(self):
        """重置窗宽窗位到默认值"""
        self.window_width = 400
        self.window_level = 40
        self.ww_slider.setValue(self.window_width)
        self.wl_slider.setValue(self.window_level)
        self.ww_value.setText(f'{self.window_width}')
        self.wl_value.setText(f'{self.window_level}')
        
        # 如果当前有显示的图像，重新显示
        if self.current_dicom_index >= 0 and self.original_pixel_array is not None:
            self.display_dicom_image(self.dicom_files[self.current_dicom_index])
            
    def prev_image(self):
        """显示上一张图像"""
        if self.dicom_files and self.current_dicom_index > 0:
            self.dicom_list.setCurrentRow(self.current_dicom_index - 1)
            
    def next_image(self):
        """显示下一张图像"""
        if self.dicom_files and self.current_dicom_index < len(self.dicom_files) - 1:
            self.dicom_list.setCurrentRow(self.current_dicom_index + 1)
            
    def apply_angle(self):
        """应用当前角度到当前子文件夹"""
        if self.current_subfolder_index < 0:
            QMessageBox.warning(self, '警告', '请先选择一个子文件夹')
            return
            
        subfolder_path = self.subfolders[self.current_subfolder_index]
        rel_path = os.path.relpath(subfolder_path, self.src_folder)
        
        # 记录角度
        self.angle_records[rel_path] = self.current_angle
        QMessageBox.information(self, '成功', f'已为 {rel_path} 设置角度 {self.current_angle}°')
        
    def save_to_excel(self):
        """保存角度记录到Excel文件"""
        if not self.dst_folder:
            QMessageBox.warning(self, '警告', '请先选择目标文件夹')
            return
            
        if not self.angle_records:
            QMessageBox.warning(self, '警告', '没有角度记录可保存')
            return
            
        # 创建Excel文件路径
        excel_path = os.path.join(self.dst_folder, 'angle_records.xlsx')
        
        try:
            # 尝试加载现有文件
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
                ws = wb.active
            else:
                # 创建新文件
                wb = Workbook()
                ws = wb.active
                ws.append(['子文件夹路径', '旋转角度'])
                
            # 添加或更新记录
            for subfolder, angle in self.angle_records.items():
                # 检查是否已存在该记录
                found = False
                for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过标题）
                    if ws.cell(row=row, column=1).value == subfolder:
                        ws.cell(row=row, column=2).value = angle
                        found = True
                        break
                        
                if not found:
                    ws.append([subfolder, angle])
                    
            # 保存文件
            wb.save(excel_path)
            QMessageBox.information(self, '成功', f'角度记录已保存到 {excel_path}')
            
        except Exception as e:
            QMessageBox.warning(self, '错误', f'保存Excel文件时出错: {str(e)}')
            
def main():
    app = QApplication(sys.argv)
    viewer = DicomViewer()
    viewer.show()
    sys.exit(app.exec_())
    
if __name__ == '__main__':
    main()